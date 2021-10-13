#-*- coding:utf-8 -*-
import re
import os
import sys
import glob
from openpyxl import load_workbook

def create_tokens_for_localization(cs_filename):
    match = []
    applied_cs_code = ''
    
    for cs_file in glob.glob(cs_filename + "/*.cs"):
        applied_cs_code = ''
        with open(cs_file, 'r', encoding='utf-8') as file:
            while True:
                line = file.read()
                if not line:
                    break

                tokens = re.findall('"!@(\w+)"', ''.join(line))

                for items in tokens:
                    match.append(items)

                line = re.sub('"!@(\w+)"', r'Lang.Msgs.\1', line, flags=re.MULTILINE)
                applied_cs_code = line

        with open(cs_file, 'w', encoding='utf-8') as file:
            file.write(applied_cs_code)
            file.close()
    
    match = list(set(match))
    return match

def add_new_tokens_to_xlsx(xlsx_filename, list_of_new_tokens):
    if not os.path.exists(xlsx_filename):
        print('no .xlsx file')
        return

    worksheet = load_workbook(xlsx_filename)["Msgs"]

    tokens_from_xlsx = []
    for cell in worksheet['A']:
        tokens_from_xlsx.append(cell.value)

    tokens_to_be_added = list(set(list_of_new_tokens) - set(tokens_from_xlsx))
            
    print("\n[[[ tokens from excel ]]] : ")
    print(tokens_from_xlsx)
    
    print("\n[[[ tokens to be added ]]] : ")
    print(tokens_to_be_added)

    for i in range(len(tokens_to_be_added)):
        worksheet.cell(len(tokens_from_xlsx) + 1 + i, 1, value = tokens_to_be_added[i])

    workbook.save(xlsx_filename)

if __name__ == '__main__':
    if len(sys.argv) > 1:
        match = create_tokens_for_localization(sys.argv[1])
        print("\nregex applied\n")
        add_new_tokens_to_xlsx(sys.argv[2], match)
    else:
        print("Usage : python3 CSharpener.py CS_FOLDERNAME FILENAME.xlsx")