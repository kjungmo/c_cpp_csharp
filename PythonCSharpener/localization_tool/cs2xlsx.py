#-*- coding:utf-8 -*-
import re
import os
import sys
import glob
from openpyxl import load_workbook

def create_tokens_for_localization(cs_filename):
    matched_tokens = []
    applied_cs_code = ''
    
    for cs_file in glob.glob(cs_filename + "/*.cs"):
        applied_cs_code = ''
        with open(cs_file, 'r', encoding='utf-8') as file:
            cs_code = file.read()
            tokens = re.findall('"!@(\w+)"', ''.join(cs_code))
            for items in tokens:
                matched_tokens.append(items)

            cs_code = re.sub('"!@(\w+)"', r'Lang.Msgs.\1', cs_code, flags=re.MULTILINE)
            applied_cs_code = cs_code

        with open(cs_file, 'w', encoding='utf-8') as file:
            file.write(applied_cs_code)
    
    matched_tokens = list(set(matched_tokens))
    return matched_tokens

def add_new_tokens_to_xlsx(xlsx_filename, list_of_new_tokens):
    if not os.path.exists(xlsx_filename):
        print('no .xlsx file')
        return
    workbook = load_workbook(xlsx_filename)
    worksheet = workbook["Msgs"]

    tokens_from_xlsx = []
    for cell in worksheet['A']:
        tokens_from_xlsx.append(cell.value)

    tokens_to_be_added = list(set(list_of_new_tokens) - set(tokens_from_xlsx))
            
    num_of_existing_tokens = len(tokens_from_xlsx)
    for i in range(len(tokens_to_be_added)):
        worksheet.cell(num_of_existing_tokens + 1 + i, 1, value = tokens_to_be_added[i])

    workbook.save(xlsx_filename)
    return tokens_to_be_added

if __name__ == '__main__':
    if len(sys.argv) > 1:
        match = create_tokens_for_localization(sys.argv[1])
        print("\nregex applied\n")
        added = add_new_tokens_to_xlsx(sys.argv[2], match)
        print("\n[[[ New tokens ]]] : ")
        print(added)
    else:
        print("Usage : python3 CSharpener.py CS_FOLDERNAME FILENAME.xlsx")