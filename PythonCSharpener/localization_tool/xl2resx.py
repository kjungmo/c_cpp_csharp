from collections import OrderedDict
import subprocess
import sys
import os
from shutil import copyfile
from openpyxl import load_workbook 


def convert_xlsx_to_resx(filename):
    if not os.path.exists(filename):
        print('the input file does not exist')
        return

    for ws in load_workbook(filename):
        res_dict = OrderedDict({c[0]: c[1:] for c in ws.iter_cols(values_only=True) if c[0] is not None})
        keys = list(res_dict.keys())

        for k in keys[1:]:
            with open('tmp.txt', 'w') as f:
                f.writelines([f'{n}={v}\n' for n, v in zip(res_dict['name'], res_dict[k]) if n is not None])

            subprocess.run(['resgen', 'tmp.txt', f'/root/workspace/localization_tool/out/{ws.title}.{k}.resx'])

        copyfile(f'/root/workspace/localization_tool/out/{ws.title}.{keys[1]}.resx', f'/root/workspace/localization_tool/out/{ws.title}.resx')

    os.remove('tmp.txt')


if __name__ == '__main__':
    if len(sys.argv) > 1:
        convert_xlsx_to_resx(sys.argv[1])
    else:
        print("Usage: python3 xl2resx.py FILENAME.xlsx")
