#-*- coding:utf-8 -*-
import re
import os
import sys
from openpyxl import load_workbook

def tokenize_localization_key_by_regex(cs_filename):
    if not os.path.exists(cs_filename):
        print('no .cs file')
        return

    match = []
    appliedCode = ''
    
    with open(cs_filename, 'r', encoding='utf-8') as file:
        while True:
            line = file.read()
            if not line:
                break
            match = re.findall('"!@(\w+)"', ''.join(line))
            line = re.sub('"!@(\w+)"', r'Lang.Msgs.\1', line, flags=re.MULTILINE)
            appliedCode = line
            
    with open(cs_filename, 'w', encoding='utf-8') as file:
        file.write(appliedCode)
        file.close()

    print("regex applied\n")
    return match

def add_new_tokens_to_xlsx(xlsx_filename, list_of_tokens):
    if not os.path.exists(xlsx_filename):
        print('no .xlsx file')
        return

    tokens_from_cs = list_of_tokens

    workbook = load_workbook(xlsx_filename)
    worksheet = workbook["Msgs"]

    tokens_from_xlsx = []
    col_a = worksheet['A']
    for cell in col_a:
        tokens_from_xlsx.append(cell.value)

    set_of_tokens = set(tokens_from_xlsx)
    tokens_to_be_added = [x for x in tokens_from_cs if x not in set_of_tokens]

    for i in range(len(tokens_to_be_added)):
        worksheet.cell(len(tokens_from_xlsx) + 1 + i, 1, value = tokens_to_be_added[i]) # row , col 

    workbook.save(xlsx_filename)

if __name__ == '__main__':
    if len(sys.argv) > 1:
        match = tokenize_localization_key_by_regex(sys.argv[1])
        add_new_tokens_to_xlsx(sys.argv[2], match)
    else:
        print("Usage : python3 CSharpener.py cs_filename.cs xlsx_filename.xlsx")